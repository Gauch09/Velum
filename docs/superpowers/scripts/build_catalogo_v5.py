# -*- coding: utf-8 -*-
import re, shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

SRC = "Cotizador VELUM - Validacion Modelo v4.xlsx"
DST = "Cotizador VELUM - Validacion Modelo v5.xlsx"
shutil.copyfile(SRC, DST)

def familia(mat):
    if mat.startswith("Galv"): return "Acero Galvanizado"
    if mat.startswith("Steel"): return "Acero"
    if mat.startswith("Alum"): return "Aluminio"
    if mat.startswith("Lux"): return "Luxsteel"
    if mat.startswith("Bond"): return "Al. Compuesto"
    if mat.startswith("Inox"): return "Inox"
    if mat.startswith("MDF"): return "MDF"
    return None

def espesor(mat):
    m = re.search(r"(\d+(?:,\d+)?)\s*mm", mat)
    return float(m.group(1).replace(",", ".")) if m else None

ACTIVE_FAMS = {"Acero Galvanizado", "Aluminio", "Luxsteel", "Al. Compuesto"}

def norm(mat):
    return mat.replace("Luxidier", "Luxsteel").strip()

def S(s):
    return s.split("|")

MS_SHEET_AZ = S("Steel 0,7mm|Galv 0,7mm|Galv 0,9mm|Luxidier 0,6mm|Alum 1mm")
MS_SHEET_A  = S("Steel 0,7mm|Galv 0,7mm|Galv 0,9mm|Luxidier 0,6mm|Alum 1mm|Bond 4mm")
MS_CLAD_AZ  = S("Steel 0,7mm|Galv 0,7mm|Galv 0,9mm|Luxidier 0,6mm|Alum 1mm")
MDF4        = S("MDF.Cedro 16mm|MDF.Melamina 19mm|MDF.Mel. 18mm H3730|MDF.Mel. 18mm W1100")

DATA = {
 "Sheet": {
   "StandardFlat": ("B-Design|Custom-Design", S("Galv 0,7mm|Galv 1,2mm|Galv 1,6mm|Galv 2mm|Galv 3,2mm|Steel 1,6mm|Steel 2mm|Luxidier 0,6mm|Alum 1,2mm")),
   "CustomFlat": ("A|B|C|D|Custom|Pattern|HalfTone", S("Alum 1,5mm|Galv 0,7mm|Galv 0,9mm|Galv 1,2mm|Galv 1,6mm|Steel 1,6mm|Steel 2mm|Steel 2,5mm|Luxidier 0,6mm|Galv 0,5mm")),
   "CustomPanel": ("A|B|C|D|Custom|Smooth|Micro|Pattern|HalfTone", S("Alum 1,5mm|Bond 4mm|Galv 0,7mm|Galv 0,9mm|Galv 1,2mm|Galv 1,6mm|Galv 3,2mm|Steel 0,5mm|Steel 1,6mm|Steel 2mm|Luxidier 0,6mm|Galv 0,5mm")),
   "Triangle": ("Estandar Sheet", S("Alum 1,5mm|Galv 0,7mm|Galv 0,9mm|Galv 1,2mm|Galv 1,6mm|Steel 1,6mm|Steel 2mm|Steel 2,5mm|Luxidier 0,6mm|Galv 0,5mm")),
   "Square": ("Estandar Sheet", S("Luxidier 0,6mm")),
   "MultiSlim.A": ("Smooth|Pattern55|Pattern70|MSHalfTone|Micro|Bond 4mm", MS_SHEET_A),
   "MultiSlim.C": ("Smooth|Pattern55|Pattern70|MSHalfTone|Micro|Bond 4mm", MS_SHEET_AZ),
   "MultiSlim.D": ("idem", MS_SHEET_AZ), "MultiSlim.E": ("idem", MS_SHEET_AZ),
   "MultiSlim.I": ("idem", MS_SHEET_AZ), "MultiSlim.L": ("idem", MS_SHEET_AZ),
   "MultiSlim.O": ("idem", MS_SHEET_AZ), "MultiSlim.V": ("idem", MS_SHEET_AZ),
   "MultiSlim.W": ("idem", MS_SHEET_AZ), "MultiSlim.U": ("idem", MS_SHEET_AZ),
 },
 "Skin": {
   "Standard": ("Smooth|Pattern|HalfTone|Custom|Scales", S("Alum 1,5mm|Alum 2,5mm|Alum 2mm|Galv 0,7mm|Galv 1,2mm|Galv 1,6mm|Galv 2mm|Galv 3,2mm")),
   "Prism": ("Smooth|Pattern|HalfTone|Custom|Scales", S("Alum 1,5mm|Galv 0,5mm|Galv 1,2mm|Steel 1,2mm|Luxidier 0,6mm|Luxidier 0,7mm|Steel 0,9mm")),
   "Lattice": ("Smooth|Pattern|Custom", S("Alum 1mm|Galv 0,7mm|Galv 0,9mm|Galv 2mm|Galv 3,2mm|Luxidier 0,6mm|Luxidier 0,7mm")),
   "Crossbar": ("Smooth|Pattern|Custom", S("Alum 1mm|Galv 0,5mm|Galv 0,7mm|Luxidier 0,5mm|Luxidier 0,6mm|Steel 0,5mm|Galv 2mm")),
   "Expanded": ("A-Diamond|B-Diamond|C-Diamond|D-Diamond", S("Alum 1,5mm|Alum 2,5mm|Alum 2mm|Galv 1,2mm|Galv 1,6mm|Steel 1,2mm|Steel 1,6mm|Galv 2mm")),
   "Composite": ("SmoothBond|PatternBond|HalfToneBond|CustomBond|ScalesBond", S("Bond 4mm|Bond 5mm")),
   "Cinetik": ("Rectangle|Rombo", S("Alum 1,5mm|Alum 1mm|Galv 0,9mm|Galv 1,2mm|Steel 0,9mm|Steel 1,2mm")),
 },
 "Skin.Rail": {
   "Standard": ("Smooth", S("Alum 1,5mm|Alum 1mm|Galv 0,9mm|Galv 1,2mm|Steel 0,9mm|Steel 1,2mm|Galv 0,7mm")),
   "MultiSlim.A": ("Smooth|Pattern55|Pattern70|MSHalfTone|Micro|Bond 4mm", S("Steel 0,7mm|Galv 0,7mm|Galv 0,9mm|Luxidier 0,6mm|Alum 1mm|Bond 4mm")),
   "MultiSlim.C": ("idem", MS_CLAD_AZ), "MultiSlim.D": ("idem", MS_CLAD_AZ),
   "MultiSlim.E": ("idem", MS_CLAD_AZ), "MultiSlim.I": ("idem", MS_CLAD_AZ),
   "MultiSlim.L": ("idem", MS_CLAD_AZ), "MultiSlim.O": ("idem", MS_CLAD_AZ),
   "MultiSlim.V": ("idem", MS_CLAD_AZ), "MultiSlim.W": ("idem", MS_CLAD_AZ),
   "MultiSlim.U": ("idem", MS_CLAD_AZ),
   "Lattice": ("Smooth|Pattern|Custom", S("Steel 0,7mm|Galv 0,7mm|Luxidier 0,6mm|Galv 2mm|Galv 3,2mm|Alum 1mm")),
 },
 "Clad": {
   "Triangle": ("Smooth|Pattern|HalfTone|Custom|Scales", S("Steel 0,5mm|Galv 0,5mm|Luxidier 0,5mm|Luxidier 0,6mm|Galv 0,9mm")),
   "Square": ("idem", S("Steel 0,5mm|Galv 0,5mm|Luxidier 0,5mm|Luxidier 0,6mm|Galv 0,9mm")),
   "MiniWaves": ("idem", S("Steel 0,5mm|Galv 0,5mm|Luxidier 0,5mm|Luxidier 0,6mm|Galv 0,9mm")),
   "Panels": ("idem", S("Steel 0,5mm|Galv 0,5mm|Luxidier 0,5mm|Luxidier 0,6mm|Galv 0,9mm")),
   "PlyWood": ("Smooth|Dots|Slots|Custom", S("MDF.Cedro 16mm|MDF.Melamina 19mm|Bond 3mm|Bond 4mm")),
   "MultiSlim.A": ("Smooth|Pattern55|Pattern70|MSHalfTone|Micro|Bond 4mm", MS_CLAD_AZ),
   "MultiSlim.C": ("idem", MS_CLAD_AZ), "MultiSlim.D": ("idem", MS_CLAD_AZ),
   "MultiSlim.E": ("idem", MS_CLAD_AZ), "MultiSlim.I": ("idem", MS_CLAD_AZ),
   "MultiSlim.L": ("idem", MS_CLAD_AZ), "MultiSlim.O": ("idem", MS_CLAD_AZ),
   "MultiSlim.V": ("idem", MS_CLAD_AZ), "MultiSlim.W": ("idem", MS_CLAD_AZ),
   "MultiSlim.U": ("idem", MS_CLAD_AZ),
   "Lattice": ("Smooth|Pattern|Custom", S("Steel 0,7mm|Galv 0,7mm|Luxidier 0,6mm|Alum 1mm|Galv 0,5mm")),
 },
 "SunShield": {
   "Prisma": ("Smooth|Pattern|HalfTone|Scales|Custom", S("Steel 0,5mm|Galv 0,5mm|Luxidier 0,6mm|Alum 1,5mm|Bond 4mm")),
   "Diamond": ("Smooth|Pattern|HalfTone|Scales|Custom", S("Steel 0,5mm|Galv 0,5mm|Luxidier 0,6mm|Alum 1,5mm")),
   "Oval": ("idem", S("Steel 0,5mm|Galv 0,5mm|Luxidier 0,6mm")),
   "Drop": ("idem", S("Steel 0,5mm|Galv 0,5mm|Luxidier 0,6mm")),
   "Alloy": ("Smooth", S("Steel 0,7mm|Galv 0,7mm|Alum 1,5mm")),
   "Square": ("Smooth", S("Steel 0,7mm|Galv 0,7mm|Galv 0,5mm|Luxidier 0,6mm|Alum 1mm")),
 },
 "SkyCap": {
   "SkyPanel": ("Smooth|Pattern|HalfTone|Scales|Custom", S("Steel 0,5mm|Galv 0,5mm|Galv 0,7mm|Luxidier 0,6mm|Bond 3mm|Bond 4mm")),
   "SkyBars": ("Smooth|Pattern|Dots", S("Steel 0,5mm|Galv 0,5mm|Luxidier 0,6mm|MDF.Melamina 19mm")),
   "SkyPlyWood": ("Smooth|Dots|Slots", MDF4),
 },
 "FullCustom": {
   "A_Tipe": ("Smooth|Pattern|HalfTone|Custom|Scales", S("Steel 0,5mm|Steel 0,7mm|Steel 0,9mm|Steel 1,2mm|Steel 1,6mm|Galv 0,5mm|Galv 0,7mm|Galv 0,9mm|Galv 1,2mm|Galv 1,6mm|Alum 0,5mm|Alum 1mm|Alum 2mm|Alum 2,5mm|Luxidier 0,5mm|Luxidier 0,6mm|Luxidier 0,7mm|Inox 0,5mm|Inox 0,7mm|Inox 0,9mm|Inox 1,2mm|Inox 1,5mm|Bond 3mm|Bond 4mm|Bond 5mm|MDF.Cedro 16mm|MDF.Melamina 19mm|MDF.Mel. 18mm H3730|MDF.Mel. 18mm W1100")),
 },
 "Organic": {
   "StandardWood": ("Smooth|Dots|Slots|Custom", MDF4),
   "CustomWood": ("idem", MDF4),
   "ChipBoard": ("Smooth|Dots|Slots", MDF4),
   "PlyWood": ("Smooth|Dots|Slots", MDF4),
 },
 "Frame": {
   "Expanded": ("Smooth|Pattern|HalfTone|Custom|Scales", S("Alum 1,5mm|Alum 2mm|Alum 2,5mm|Galv 0,7mm|Galv 1,2mm|Galv 1,6mm|Galv 2mm|Galv 3,2mm|Steel 1,6mm|Steel 2mm")),
 },
}
for t in ("B_Tipe", "C_Tipe", "D_Tipe"):
    DATA["FullCustom"][t] = (DATA["FullCustom"]["A_Tipe"][0], list(DATA["FullCustom"]["A_Tipe"][1]))

rows = []
for sis, prods in DATA.items():
    for prod, (dis, mats) in prods.items():
        clean = []
        for raw in mats:
            mat = norm(raw)
            fam = familia(mat); esp = espesor(mat)
            if fam is None or esp is None or mat in clean:
                continue
            clean.append(mat)
            activo = "Si" if (sis == "Skin" and fam in ACTIVE_FAMS) else "No"
            rows.append([sis, prod, dis, mat, fam, esp, activo])

skin_lists = {}
for sis, prod, dis, mat, fam, esp, activo in rows:
    if activo == "Si":
        skin_lists.setdefault(prod, [])
        if mat not in skin_lists[prod]:
            skin_lists[prod].append(mat)
skin_products = list(skin_lists.keys())

wb = openpyxl.load_workbook(DST)
HEAD = Font(bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="2F5496")
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
GREEN = PatternFill("solid", fgColor="E2EFDA")

if "Catalogo Variantes" in wb.sheetnames:
    del wb["Catalogo Variantes"]
cat = wb.create_sheet("Catalogo Variantes")
heads = ["Sistema", "Producto", "Disenos", "Material", "Familia", "Espesor_mm", "Activo_Simulador"]
for j, h in enumerate(heads, start=1):
    c = cat.cell(row=1, column=j, value=h); c.font = HEAD; c.fill = HFILL; c.border = BORD
for i, r in enumerate(rows, start=2):
    for j, v in enumerate(r, start=1):
        c = cat.cell(row=i, column=j, value=v); c.border = BORD
        if r[6] == "Si":
            c.fill = GREEN
for j, w in enumerate([12, 16, 40, 18, 18, 12, 16], start=1):
    cat.column_dimensions[get_column_letter(j)].width = w
cat.freeze_panes = "A2"

if "Listas" in wb.sheetnames:
    del wb["Listas"]
lst = wb.create_sheet("Listas")
lst.cell(row=1, column=1, value="Productos Skin").font = Font(bold=True)
for i, p in enumerate(skin_products, start=2):
    lst.cell(row=i, column=1, value=p)
prod_ref = "Listas!$A$2:$A$%d" % (1 + len(skin_products))
col = 3
named = {}
for p in skin_products:
    mats = skin_lists[p]
    letter = get_column_letter(col)
    lst.cell(row=1, column=col, value=p).font = Font(bold=True)
    for i, m in enumerate(mats, start=2):
        lst.cell(row=i, column=col, value=m)
    named["mat_" + p] = "Listas!$%s$2:$%s$%d" % (letter, letter, 1 + len(mats))
    lst.column_dimensions[letter].width = 16
    col += 1
lst.column_dimensions["A"].width = 16

def set_name(nm, ref):
    if nm in wb.defined_names:
        del wb.defined_names[nm]
    wb.defined_names[nm] = DefinedName(nm, attr_text=ref)

set_name("sys_Skin", prod_ref)
for nm, ref in named.items():
    set_name(nm, ref)

sim = wb["Simulador"]
sim.data_validations.dataValidation = []
dv_sys = DataValidation(type="list", formula1='"Skin"', allow_blank=False)
dv_sys.add(sim["B7"]); sim.add_data_validation(dv_sys)
dv_prod = DataValidation(type="list", formula1="sys_Skin", allow_blank=False)
dv_prod.add(sim["B8"]); sim.add_data_validation(dv_prod)
dv_mat = DataValidation(type="list", formula1='INDIRECT("mat_"&$B$8)', allow_blank=False)
dv_mat.add(sim["B10"]); sim.add_data_validation(dv_mat)

sim["B7"] = "Skin"
sim["B8"] = "Standard"
sim["B10"] = "Alum 1,5mm"
sim["A11"] = "Espesor panel (mm) auto"
sim["B11"] = "=IFERROR(VLOOKUP(B10,'Catalogo Variantes'!$D:$F,3,FALSE),0)"
sim["A15"] = "Familia (auto)"
sim["B15"] = "=IFERROR(VLOOKUP(B10,'Catalogo Variantes'!$D:$F,2,FALSE),\"\")"
sim["B17"] = "=IFERROR(VLOOKUP(B15,Parametros!$A$22:$D$27,3,FALSE),0)"
sim["B18"] = "=IFERROR(VLOOKUP(B15,Parametros!$A$22:$D$27,2,FALSE),0)"
for addr in ("B9", "B11", "B15", "B17", "B18"):
    sim[addr].fill = GREEN

wb.save(DST)
print("OK saved", DST)
print("catalog rows:", len(rows))
print("Skin products:", skin_products)
print("named ranges:", ["sys_Skin"] + list(named.keys()))
for p in skin_products:
    print("  ", p, "->", skin_lists[p])

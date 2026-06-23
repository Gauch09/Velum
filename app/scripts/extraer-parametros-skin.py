# Lee la planilla viva y emite INSERTs SQL para los parametros del motor Skin.
# Tablas: MaterialFamilia, DisenoKp, MaterialVariante, ParametroCosteo (escalares Skin).
# Patron F0: celdas data_only, raise en None, cast a float.
# Uso: python extraer-parametros-skin.py [ruta_xlsx]
import sys
from openpyxl import load_workbook

RUTA = sys.argv[1] if len(sys.argv) > 1 else "Cotizador VELUM - Validacion Modelo v7.xlsx"
wb = load_workbook(RUTA, data_only=True)

def cell(hoja, ref):
    return wb[hoja][ref].value

# ---------------------------------------------------------------------------
# 1. MaterialFamilia — Parametros A22:E27
#    col A=nombre, B=densidad, C=precioTon, D=nota, E=precioM2 (None si vacio)
# ---------------------------------------------------------------------------
print("-- MaterialFamilia")
FAMILIAS_ROWS = range(22, 28)  # rows 22..27
for fila in FAMILIAS_ROWS:
    nombre   = wb["Parametros"][f"A{fila}"].value
    densidad = wb["Parametros"][f"B{fila}"].value
    precio_ton = wb["Parametros"][f"C{fila}"].value
    precio_m2_raw = wb["Parametros"][f"E{fila}"].value

    if nombre is None:
        raise ValueError(f"MaterialFamilia fila {fila}: nombre vacio")
    if densidad is None:
        raise ValueError(f"MaterialFamilia fila {fila} ({nombre}): densidad vacia")
    if precio_ton is None:
        raise ValueError(f"MaterialFamilia fila {fila} ({nombre}): precioTon vacio")

    densidad   = float(densidad)
    precio_ton = float(precio_ton)
    precio_m2  = float(precio_m2_raw) if precio_m2_raw is not None else 0.0

    nombre_safe = nombre.replace("'", "''")
    print(
        f"INSERT INTO \"MaterialFamilia\" (id, nombre, densidad, \"precioTon\", \"precioM2\", \"actualizadoEn\") "
        f"VALUES (gen_random_uuid()::text, '{nombre_safe}', {densidad}, {precio_ton}, {precio_m2}, now()) "
        f"ON CONFLICT (nombre) DO UPDATE SET densidad = EXCLUDED.densidad, "
        f"\"precioTon\" = EXCLUDED.\"precioTon\", \"precioM2\" = EXCLUDED.\"precioM2\", "
        f"\"actualizadoEn\" = now();"
    )

# ---------------------------------------------------------------------------
# 2. DisenoKp — Factores Kp A3:B12
# ---------------------------------------------------------------------------
print()
print("-- DisenoKp")
ws_kp = wb["Factores Kp"]
for row in ws_kp.iter_rows(min_row=3, max_row=12, min_col=1, max_col=2, values_only=True):
    diseno, kp = row
    if diseno is None or kp is None:
        continue
    kp = float(kp)
    diseno_safe = str(diseno).replace("'", "''")
    print(
        f"INSERT INTO \"DisenoKp\" (id, diseno, kp, \"actualizadoEn\") "
        f"VALUES (gen_random_uuid()::text, '{diseno_safe}', {kp}, now()) "
        f"ON CONFLICT (diseno) DO UPDATE SET kp = EXCLUDED.kp, \"actualizadoEn\" = now();"
    )

# ---------------------------------------------------------------------------
# 3. MaterialVariante — Catalogo Variantes cols D(material)/E(familia)/F(espesor)
#    Dedup por material (primera ocurrencia gana), skip si D es None.
#    Solo incluir familias conocidas en MaterialFamilia (Skin). Omitir MDF y cualquier otra.
# ---------------------------------------------------------------------------
FAMILIAS_SKIN = {
    'Acero Galvanizado', 'Acero', 'Aluminio', 'Inox', 'Al. Compuesto', 'Luxsteel'
}

print()
print("-- MaterialVariante")
ws_cat = wb["Catalogo Variantes"]
seen_mat = set()
for row in ws_cat.iter_rows(min_row=2, min_col=4, max_col=6, values_only=True):
    mat, fam, esp = row
    if mat is None:
        continue
    if mat in seen_mat:
        continue
    seen_mat.add(mat)
    if fam is None:
        raise ValueError(f"MaterialVariante '{mat}': familia vacia")
    if str(fam) not in FAMILIAS_SKIN:
        # Familia no es un material Skin (ej. MDF) — omitir del seed
        continue
    if esp is None:
        raise ValueError(f"MaterialVariante '{mat}': espesorMm vacio")
    esp = float(esp)
    mat_safe = str(mat).replace("'", "''")
    fam_safe = str(fam).replace("'", "''")
    print(
        f"INSERT INTO \"MaterialVariante\" (id, material, familia, \"espesorMm\", \"actualizadoEn\") "
        f"VALUES (gen_random_uuid()::text, '{mat_safe}', '{fam_safe}', {esp}, now()) "
        f"ON CONFLICT (material) DO UPDATE SET familia = EXCLUDED.familia, "
        f"\"espesorMm\" = EXCLUDED.\"espesorMm\", \"actualizadoEn\" = now();"
    )

# ---------------------------------------------------------------------------
# 4. ParametroCosteo — escalares Skin
# ---------------------------------------------------------------------------
SCALARS = {
    # clave: (hoja, celda, unidad, descripcion)  — valor None = literal abajo
    "skin_costilla_area":      ("Simulador",        "B19", "m2",    "Area costilla Skin"),
    "skin_costilla_espesor":   ("Simulador",        "B20", "mm",    "Espesor costilla Skin"),
    "skin_mensula_area":       ("Simulador",        "B22", "m2",    "Area mensula Skin"),
    "skin_mensula_espesor":    ("Simulador",        "B23", "mm",    "Espesor mensula Skin"),
    "pintura_polvo":           ("Parametros",       "H4",  "u$d/kg","Precio pintura polvo"),
    "pintura_cobertura":       ("Parametros",       "H5",  "m2/kg", "Cobertura pintura polvo"),
    "pintura_sobreaplic":      ("Parametros",       "H6",  "-",     "Factor sobre-aplicacion pintura"),
    "pintura_horneada_costo":  ("Parametros",       "H8",  "u$d/pz","Costo horneado por pieza"),
    "pintura_horneada_piezas": ("Parametros",       "H9",  "pz",    "Piezas por horneada"),
    "fijacion_broca":          ("Insumos Fijacion", "B7",  "u$d",   "Costo broca por unidad"),
    "fijacion_autoperf":       ("Insumos Fijacion", "B5",  "u$d",   "Costo tornillo autoperforante"),
    "acm_area_placa":          ("Parametros",       "H17", "m2",    "Area placa ACM"),
    "acm_fab_placa":           ("Parametros",       "H18", "u$d/pz","Fab costo placa ACM"),
    "acm_acc_panel":           ("Parametros",       "H21", "pz",    "Accesorios por panel ACM"),
    "acm_acc_costo":           ("Parametros",       "H20", "u$d",   "Costo accesorio ACM"),
    "galv_densidad":           ("Parametros",       "B22", "kg/m3", "Densidad Acero Galvanizado"),
    "galv_precio_ton":         ("Parametros",       "C22", "u$d/t", "Precio ton Acero Galvanizado"),
}

# Literales fijos (no estan en celda propia)
LITERALS = {
    "skin_empalme_area":    (0.0387,  "m2",   "Area empalme costilla Skin"),
    "skin_empalme_espesor": (1.6,     "mm",   "Espesor empalme costilla Skin"),
}

print()
print("-- ParametroCosteo (escalares Skin)")
for clave, (hoja, ref, unidad, desc) in SCALARS.items():
    v = cell(hoja, ref)
    if v is None:
        raise ValueError(f"celda vacia: {clave} ({hoja}!{ref})")
    v = float(v)
    desc_safe = desc.replace("'", "''")
    print(
        f"INSERT INTO \"ParametroCosteo\" (id, clave, valor, unidad, descripcion, \"actualizadoEn\") "
        f"VALUES (gen_random_uuid()::text, '{clave}', {v}, '{unidad}', '{desc_safe}', now()) "
        f"ON CONFLICT (clave) DO UPDATE SET valor = EXCLUDED.valor, \"actualizadoEn\" = now();"
    )

for clave, (v, unidad, desc) in LITERALS.items():
    desc_safe = desc.replace("'", "''")
    print(
        f"INSERT INTO \"ParametroCosteo\" (id, clave, valor, unidad, descripcion, \"actualizadoEn\") "
        f"VALUES (gen_random_uuid()::text, '{clave}', {v}, '{unidad}', '{desc_safe}', now()) "
        f"ON CONFLICT (clave) DO UPDATE SET valor = EXCLUDED.valor, \"actualizadoEn\" = now();"
    )
